# Com momentum, inicialização aleatória, sigmoide na oculta e linear na Saída, gráfico corrigido

import torch
import torch.nn as nn
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook
import random
import time

# Configurações para reprodutibilidade
torch.manual_seed(42)
np.random.seed(42)
random.seed(42)

# Criar pastas se não existirem
os.makedirs('Resultados', exist_ok=True)

# Definir a rede neural conforme especificações do projeto
class MLP(nn.Module):
    def __init__(self):
        super(MLP, self).__init__()
        # Arquitetura: 3 entradas -> 10 neurônios ocultos -> 1 saída
        self.hidden = nn.Linear(3, 10)
        self.output = nn.Linear(10, 1)
        
        # Função de ativação sigmoide
        self.sigmoid = nn.Sigmoid()
        
        # Inicialização com pesos aleatórios pequenos (método clássico para sigmoides)
        self._initialize_weights()
    
    def _initialize_weights(self):
        # Inicialização clássica para redes com sigmoides
        for layer in [self.hidden, self.output]:
            nn.init.uniform_(layer.weight, -0.5, 0.5)
            nn.init.zeros_(layer.bias)
    
    def forward(self, x):
        # Camada oculta com sigmoide
        x = self.hidden(x)
        x = self.sigmoid(x)
        
        # Camada de saída linear (sem ativação)
        x = self.output(x)
        
        return x

# Função de treinamento seguindo especificações do projeto
def treinar_rede(X_train, y_train, learning_rate=0.1, max_epochs=3000000, target_error=1e-6, train_num=1):
    model = MLP()
    
    # SGD com momentum para melhor convergência
    optimizer = torch.optim.SGD(model.parameters(), lr=learning_rate, momentum=0.9)
    
    # Função de perda: erro quadrático médio
    criterion = nn.MSELoss()
    
    errors = []
    epoch = 0
    
    model.train()
    
    print(f"  Iniciando Treinamento {train_num}...")
    
    while epoch < max_epochs:
        # Forward pass
        outputs = model(X_train)
        loss = criterion(outputs, y_train)
        
        # Backward pass
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
        
        current_loss = loss.item()
        # Armazenar todos os erros para gráficos precisos
        errors.append(current_loss)
        
        epoch += 1
        
        # Exibir progresso a cada 10.000 épocas
        if epoch % 10000 == 0:
            print(f"    Treinamento {train_num} - Época {epoch:,}: Erro = {current_loss:.6e}")
        
        # Critério de parada
        if current_loss < target_error:
            print(f"    Treinamento {train_num} - Convergiu na época {epoch:,} com erro {current_loss:.6e}")
            break
    
    # Se não convergiu
    if epoch == max_epochs:
        print(f"    Treinamento {train_num} - Atingiu máximo de épocas ({max_epochs:,}) com erro {current_loss:.6e}")
    
    return model, errors, epoch

# Carregar dados de treinamento
print("Carregando dados de treinamento...")
try:
    df_train = pd.read_excel('Dados do projeto/DadosProjeto01RNA.xlsx', sheet_name='DadosTreinamentoRNA')
    X_train_np = df_train[['x1 ', 'x2 ', 'x3 ']].values
    y_train_np = df_train['d '].values.reshape(-1, 1)
    
    # Converter para tensores
    X_train = torch.FloatTensor(X_train_np)
    y_train = torch.FloatTensor(y_train_np)
    
    print(f"✓ Dados carregados: {X_train.shape[0]} amostras de treinamento")
    print(f"✓ Entradas: {X_train.shape[1]} variáveis (x1, x2, x3)")
    
except Exception as e:
    print(f"ERRO ao carregar dados: {e}")
    print("Certifique-se de que 'Dados do projeto/DadosProjeto01RNA.xlsx' existe")
    exit()

# Executar 5 treinamentos conforme solicitado no projeto
resultados_treinamento = []
tempo_total = time.time()

print("\n" + "="*70)
print("INICIANDO 5 TREINAMENTOS DA RNA - PROJETO ETAPA 01")
print("="*70)
print("Parâmetros:")
print("- Taxa de aprendizado (η): 0.1")
print("- Precisão (ε): 1e-6")
print("- Função de ativação: Sigmoide (todas as camadas)")
print("- Algoritmo: Backpropagation (SGD sem momentum)")
print("-"*70)

for i in range(5):
    # Reinicializar semente para cada treinamento (valores iniciais diferentes)
    torch.manual_seed(42 + i*100)
    np.random.seed(42 + i*100)
    random.seed(42 + i*100)
    
    print(f"\nTREINAMENTO {i+1}/5")
    print("-" * 50)
    tempo_inicio = time.time()
    
    # Treinar a rede
    model, errors, epochs = treinar_rede(X_train, y_train, train_num=i+1)
    
    tempo_fim = time.time()
    tempo_treinamento = tempo_fim - tempo_inicio
    
    # Calcular erro quadrático médio final
    model.eval()
    with torch.no_grad():
        outputs = model(X_train)
        mse_final = nn.MSELoss()(outputs, y_train).item()
    
    print(f"    ✓ Tempo de treinamento: {tempo_treinamento:.2f} segundos")
    print(f"    ✓ Erro final: {mse_final:.6e}")
    
    resultados_treinamento.append({
        'treinamento': i+1,
        'model': model,
        'errors': errors,
        'epochs': epochs,
        'mse': mse_final,
        'tempo': tempo_treinamento
    })

tempo_total = time.time() - tempo_total
print(f"\n{'='*70}")
print(f"RESUMO GERAL:")
print(f"Tempo total de treinamento: {tempo_total:.2f} segundos")
print(f"Média de épocas: {np.mean([r['epochs'] for r in resultados_treinamento]):.0f}")
print(f"Melhor erro: {min([r['mse'] for r in resultados_treinamento]):.6e}")
print("="*70)

# Criar Tabela 1 - Resultados dos Treinamentos
print("\nGerando Tabela 1 - Resultados dos Treinamentos...")
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Tabela 1 - Resultados Treinamento"

# Cabeçalhos
ws1['A1'] = 'Treinamento'
ws1['B1'] = 'Erro Quadrático Médio'
ws1['C1'] = 'Número Total de Épocas'

# Dados dos treinamentos
for i, resultado in enumerate(resultados_treinamento):
    ws1[f'A{i+2}'] = f"{resultado['treinamento']}º (T{resultado['treinamento']})"
    ws1[f'B{i+2}'] = f"{resultado['mse']:.2e}"
    ws1[f'C{i+2}'] = resultado['epochs']

# Salvar Tabela 1
wb1.save('Resultados/tabela1_resultados_treinamento.xlsx')
print("✓ Tabela 1 salva em: Resultados/tabela1_resultados_treinamento.xlsx")

# Identificar os dois treinamentos com maior número de épocas (conforme item 3)
sorted_results = sorted(resultados_treinamento, key=lambda x: x['epochs'], reverse=True)
top_two = sorted_results[:2]

print(f"\nTreinamentos com maior número de épocas:")
print(f"1º: T{top_two[0]['treinamento']} com {top_two[0]['epochs']} épocas")
print(f"2º: T{top_two[1]['treinamento']} com {top_two[1]['epochs']} épocas")

# Criar gráficos dos dois treinamentos com maior número de épocas
print("\nGerando gráficos de convergência...")
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 10))

# Gráfico 1 - Primeiro treinamento com mais épocas
train_num1 = top_two[0]['treinamento']
errors1 = top_two[0]['errors']
epochs1 = top_two[0]['epochs']

ax1.plot(range(1, epochs1+1), errors1, 'b-', linewidth=1)
ax1.set_xlabel('Época')
ax1.set_ylabel('Erro Quadrático Médio')
ax1.set_title(f'Convergência do Treinamento T{train_num1} - {epochs1:,} épocas')
ax1.set_yscale('log')
ax1.grid(True, alpha=0.3)
ax1.set_xlim(1, epochs1)

# Gráfico 2 - Segundo treinamento com mais épocas
train_num2 = top_two[1]['treinamento']
errors2 = top_two[1]['errors']
epochs2 = top_two[1]['epochs']

ax2.plot(range(1, epochs2+1), errors2, 'r-', linewidth=1)
ax2.set_xlabel('Época')
ax2.set_ylabel('Erro Quadrático Médio')
ax2.set_title(f'Convergência do Treinamento T{train_num2} - {epochs2:,} épocas')
ax2.set_yscale('log')
ax2.grid(True, alpha=0.3)
ax2.set_xlim(1, epochs2)

plt.tight_layout()
plt.savefig('Resultados/graficos_treinamento.png', dpi=300, bbox_inches='tight')
plt.close()
print("✓ Gráficos salvos em: Resultados/graficos_treinamento.png")

# Dados de teste fornecidos na Tabela 2 do projeto
dados_teste = [
    [0.0611, 0.2860, 0.7464, 0.4831],
    [0.5102, 0.7464, 0.0860, 0.5965],
    [0.0004, 0.6916, 0.5006, 0.5318],
    [0.9430, 0.4476, 0.2648, 0.6843],
    [0.1399, 0.1610, 0.2477, 0.2872],
    [0.6423, 0.3229, 0.8567, 0.7663],
    [0.6492, 0.0007, 0.6422, 0.5666],
    [0.1818, 0.5078, 0.9046, 0.6601],
    [0.7382, 0.2647, 0.1916, 0.5427],
    [0.3879, 0.1307, 0.8656, 0.5836],
    [0.1903, 0.6523, 0.7820, 0.6950],
    [0.8401, 0.4490, 0.2719, 0.6790],
    [0.0029, 0.3264, 0.2476, 0.2956],
    [0.7088, 0.9342, 0.2763, 0.7742],
    [0.1283, 0.1882, 0.7253, 0.4662],
    [0.8882, 0.3077, 0.8931, 0.8093],
    [0.2225, 0.9182, 0.7820, 0.7581],
    [0.1957, 0.8423, 0.3085, 0.5826],
    [0.9991, 0.5914, 0.3933, 0.7938],
    [0.2299, 0.1524, 0.7353, 0.5012]
]

# Preparar dados de teste
print("\nExecutando validação com dados de teste...")
df_teste = pd.DataFrame(dados_teste, columns=['x1', 'x2', 'x3', 'd'])
X_test = torch.FloatTensor(df_teste[['x1', 'x2', 'x3']].values)
y_test = df_teste['d'].values

# Criar Tabela 2 - Validação
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Tabela 2 - Validação"

# Cabeçalhos da Tabela 2
headers = ['Amostra', 'x1', 'x2', 'x3', 'd', 'y (T1)', 'y (T2)', 'y (T3)', 'y (T4)', 'y (T5)']
for col, header in enumerate(headers, 1):
    ws2.cell(row=1, column=col, value=header)

# Preencher dados de entrada (x1, x2, x3, d)
for i in range(20):
    ws2.cell(row=i+2, column=1, value=i+1)  # Amostra
    ws2.cell(row=i+2, column=2, value=dados_teste[i][0])  # x1
    ws2.cell(row=i+2, column=3, value=dados_teste[i][1])  # x2
    ws2.cell(row=i+2, column=4, value=dados_teste[i][2])  # x3
    ws2.cell(row=i+2, column=5, value=dados_teste[i][3])  # d (valor desejado)

# Calcular saídas para cada treinamento
erros_relativos = []
print("Calculando saídas para cada treinamento:")

for j, resultado in enumerate(resultados_treinamento):
    model = resultado['model']
    model.eval()
    
    # Calcular saídas da rede
    with torch.no_grad():
        outputs = model(X_test).numpy().flatten()
    
    # Preencher saídas na tabela
    for i in range(20):
        ws2.cell(row=i+2, column=6+j, value=f"{outputs[i]:.4f}")
    
    # Calcular erro relativo médio (%)
    erros = np.abs((y_test - outputs) / y_test) * 100
    erro_medio = np.mean(erros)
    variancia = np.var(erros)
    
    erros_relativos.append({
        'treinamento': j+1,
        'erro_medio': erro_medio,
        'variancia': variancia,
        'outputs': outputs
    })
    
    print(f"  T{j+1}: Erro relativo médio = {erro_medio:.2f}%, Variância = {variancia:.2f}%")

# Adicionar erro relativo médio e variância na tabela
ws2.cell(row=22, column=1, value='Erro relativo médio (%)')
ws2.cell(row=23, column=1, value='Variância (%)')

for j, erro_info in enumerate(erros_relativos):
    ws2.cell(row=22, column=6+j, value=f"{erro_info['erro_medio']:.2f}")
    ws2.cell(row=23, column=6+j, value=f"{erro_info['variancia']:.2f}")

# Salvar Tabela 2
wb2.save('Resultados/tabela2_validacao.xlsx')
print("✓ Tabela 2 salva em: Resultados/tabela2_validacao.xlsx")

# Determinar o melhor treinamento (menor erro relativo médio)
melhor_treinamento = min(erros_relativos, key=lambda x: x['erro_medio'])

print(f"\n{'='*70}")
print("ANÁLISE FINAL - MELHOR CONFIGURAÇÃO:")
print("="*70)
print(f"Treinamento mais adequado: T{melhor_treinamento['treinamento']}")
print(f"Erro relativo médio: {melhor_treinamento['erro_medio']:.2f}%")
print(f"Variância: {melhor_treinamento['variancia']:.2f}%")
print(f"Épocas necessárias: {resultados_treinamento[melhor_treinamento['treinamento']-1]['epochs']:,}")
print(f"Tempo de treinamento: {resultados_treinamento[melhor_treinamento['treinamento']-1]['tempo']:.2f}s")

print(f"\nRANKING DOS TREINAMENTOS (por erro relativo médio):")
ranking = sorted(erros_relativos, key=lambda x: x['erro_medio'])
for i, r in enumerate(ranking):
    print(f"{i+1}º. T{r['treinamento']}: {r['erro_medio']:.2f}% (variância: {r['variancia']:.2f}%)")

print("="*70)
print("ARQUIVOS GERADOS:")
print("✓ Resultados/tabela1_resultados_treinamento.xlsx")
print("✓ Resultados/tabela2_validacao.xlsx") 
print("✓ Resultados/graficos_treinamento.png")
print("="*70)
print("PROJETO ETAPA 01 CONCLUÍDO COM SUCESSO!")